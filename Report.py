# -*- coding: utf-8 -*-
"""
Created on Fri Feb  1 08:54:19 2019
V 6.04 (2019-04-04)
@author: ludovicraymond
"""
import datetime
import time
import os
import pandas as pd
import NordicLamSplitAnalysisFunctions as fu
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from tkinter import messagebox

SheetNRep = 'Report'

###-_-
### Update: Add to selected deco the V or W values based on the Shear equation value. 
# special Vf or Wf new var = 'shear_force', special Vr or Wr new var = 'shear_resistance', 
### ADD shear_resistance_Vf

file_path = ''
report_sheet_name = 'report'

def Report(selected_data_dico):
    '''
        Organizes the data dico information in a formated excel sheet.
    '''
    
    if selected_data_dico['shear_method_a_used']['value'] == 1:
        selected_data_dico['shear_force']['value'] = selected_data_dico['shear_force_Wf']['value']
        selected_data_dico['shear_resistance']['value'] = selected_data_dico['shear_resistance_Wf']['value']
    elif selected_data_dico['shear_method_a_used']['value'] == 0:
        selected_data_dico['shear_force']['value'] = selected_data_dico['shear_force_Vf']['value']
        selected_data_dico['shear_resistance']['value'] = selected_data_dico['shear_resistance_Vf']['value']    

    Nordic_Lam_split_analyser_general_report_template = (
            pd.read_excel('templates\\NLSA_general_report.xlsx', sheet_name=0, header=0, index_col=0))    
    lookup_list = Nordic_Lam_split_analyser_general_report_template.index

    for selected_index in lookup_list:
        try:
            Nordic_Lam_split_analyser_general_report_template.at[selected_index,'value'] = str(
                    selected_data_dico[selected_index]['value'])
        except:
            pass
        try:
            Nordic_Lam_split_analyser_general_report_template.at[selected_index,'unit'] = str(
                    selected_data_dico[selected_index]['unit'])
        except:
            pass
    Nordic_Lam_split_analyser_general_report_template = (
            Nordic_Lam_split_analyser_general_report_template.replace(to_replace='None',value=''))

    Nordic_Lam_split_analyser_general_report_template = (
            Nordic_Lam_split_analyser_general_report_template [['label','value','unit','description']])

    

    print(Nordic_Lam_split_analyser_general_report_template.columns)

#    print (Nordic_Lam_split_analyser_general_report_template.loc[:,'value'])
#    print (Nordic_Lam_split_analyser_general_report_template.loc[:,'unit'])
    print (Nordic_Lam_split_analyser_general_report_template)
    
    # Build first set of data
    df = Nordic_Lam_split_analyser_general_report_template

    #Generate File name with a time stamp
    Time=str(datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H_%M_%S'))
#    FileName = df.at['ReportType','Label']+'_'+Time+'.xlsx'
    FileName = file_path+df.at['report_type','label']+'_'+Time+'.xlsx'

    writer = pd.ExcelWriter(FileName)

    df.to_excel(writer, header=False, index=False, sheet_name=report_sheet_name)

    #Add inforation
    Notes = [('Project: ',selected_data_dico['project_name']['value']),
             ('Notes',selected_data_dico['Notes']['value']),
             ('Date: ', str(datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d'))),
             ('Designer: ', os.getlogin())] 
    df1 = pd.DataFrame(data=Notes)
    df1.to_excel(writer, index=False, header=False, sheet_name=report_sheet_name, startrow = df.shape[0] + 1)

# Add the references
#    df3 = pd.DataFrame(data=References)
#    df3.to_excel(writer, index=False, header=False, sheet_name=report_sheet_name, startcol = 1, startrow = df.shape[0] + df1.shape[0] + 2)

    #Save all the inputed data
    writer.save()







    ## Read template create new sheat with date and name etc...
    ## Replace each cell with the proper data based on lookup
    ## Add the information, project etc...
    ## Add the format
    
    
    
    return



def old_Report(Input):
    #Generates the data to be displayed in the excel spread sheet
    #Remove Notes from function input to use it elsewere as the variable Note
    ValList = []
    for i in Input:
        ValList += [i]
    #Adds the project name, the date and login to identify the project and user.
    Note = [(ValList[1]),(str(datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d'))),(os.getlogin())]
    Project = ValList[0]

    #Remove project Name and notes
    del ValList[0:2]
    #Number of items passed from the main GUI to the reports minus Fv
    MainGUIDataItems = 22
    #Remove the Fv caracteristic thats redundant
    del ValList[MainGUIDataItems]

    ASSYGUIItems = MainGUIDataItems + 9 #Main GUI + ASSY GUI
    PanelGUIItems = MainGUIDataItems + 4 #Main GUI + Panel GUI, To update with righ qty of parameters

    #Main GUI Items Units and labels
### Coding Note, When doing the imperial report, add the imperial units to the dictionary and update the dataframe builder to
### example:         if radiobutton set to Imperial then: MGuiUni += [MGuiLabUniDic[i][1]] else [MGuiLabUniDic[i][0]]


    #Set all the report Main interface parameter and output labels and units.
    MGuiLabUniDic = {'Nordic Lam Type':'Type',  'Ply':'nbr',                    'b, Width':'mm',                'h, Depth':'mm',
                     'hd, Hole diameter':'mm',  'Method':'',                    'Mf, Moment force':'Nm',        'Mr, Moment resistance':'Nm',
                     'Mrr, Reduced':'Nm',       'Mf/Mrr':'Bending',             'Shear equation used':'',       'Vf, Shear force (b)':'N',
                     'Wf, Shear force (a)':'N', 'Wr, Shear resistance (a)':'N', 'Residual Shear Res.':'N',      'Force/Resistance':'Shear',
                     'KD, Load duration':'',    'KH, system factor':'',         'Ftp: Perp. tension':'MPa',     'Ftpr, tension resistance':'MPa',
                     'Ftp/Ftpr':'P.Tension',  'Tp, Tension force':'N'}
    #Set the screw repair outpul labels and units
    SGuiLabUniDic = {'Method':'',           'Orientation':'',   'Screw Type':'',            'Screw Qty':'',
                     'Screw Length':'mm',   'Min Thread':'mm',  'Screw Resistance':'N/mm',  'Screw Offset':'mm',
                     'Tp Used':'N'}

    PGuiLabUniDic = {'Method':'',           'Panel height':'mm', 'Panel width':'mm',        '0.131" x 3.25" Nails.':'Qty'}

#    Generate the lists to build the dataframe
    MGuiUni = []
    for i in MGuiLabUniDic:
        MGuiUni += [MGuiLabUniDic[i]]
    MGuiLabel= []    
    for i in MGuiLabUniDic:
        MGuiLabel += [i]
    SGuiUni = []
    for i in SGuiLabUniDic:
        SGuiUni += [SGuiLabUniDic[i]]
    SGuiLabel= []    
    for i in SGuiLabUniDic:
        SGuiLabel += [i]  
    PGuiUni = []
    for i in PGuiLabUniDic:
        PGuiUni += [PGuiLabUniDic[i]]
    PGuiLabel= []    
    for i in PGuiLabUniDic:
        PGuiLabel += [i]      
    
    if len(ValList) > MainGUIDataItems:
        try:
            if ValList[MainGUIDataItems] == 'ASSY Screw' and len(ValList) == ASSYGUIItems:
                print('test in report function step 2')
                RepList = [ValList,MGuiUni+SGuiUni]
                df = pd.DataFrame(data=RepList)
                df.columns=MGuiLabel+SGuiLabel
                df = df.transpose()
                df.columns=['Value','Units']
            elif ValList[MainGUIDataItems] == 'Glued Panel' and len(ValList) == PanelGUIItems:
                RepList = [ValList,MGuiUni+PGuiUni]
                df = pd.DataFrame(data=RepList)
                df.columns=MGuiLabel+PGuiLabel
                df = df.transpose()
                df.columns=['Value','Units']
        except:
            print('Problem in retrieving the repair type variable or invalid output items')
    elif len(ValList) == MainGUIDataItems:
        RepList = [ValList,MGuiUni]
        df = pd.DataFrame(data=RepList)
        df.columns=MGuiLabel
        df = df.transpose()
        df.columns=['Value','Units']
    else:
        print('Invalid quantity of parameters for report.\nInput all values in app.')
        messagebox.showinfo('Missing Data','Fill up all data on the user interface')
        df = pd.DataFrame(data=[])
    
    #Generate File name with a time stamp
    Time=str(datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H_%M_%S'))
    FileName = 'NLSA_OUT_'+Time+'.xlsx'

    writer = pd.ExcelWriter(FileName)

    df.to_excel(writer, sheet_name=SheetNRep)

    #Add inforation
    df1 = pd.DataFrame(data=Note)
    df1.columns=[Project]
    df1.index=['Notes','Date','Designer']
    df1.to_excel(writer, sheet_name=SheetNRep, startrow = len(ValList)+2)

    #Add the spec used for this project
    df2 = fu.NLSpecs().transpose()
    df2.to_excel(writer, sheet_name=SheetNRep, startcol = 5)

    #Save all the inputed data
    writer.save()
    
#   Create a workbook object.
    wb = Workbook()
    wb = load_workbook(FileName)
    
#   Open up the last edited workbook sheet.    
    sheet=wb.active

#   Definition of the cells to align and merge. (Nice to have, search and find the cell ID based on value)
    ProjectCell = 'B'+str(len(ValList)+3)
    NoteCell = 'B'+str(len(ValList)+4)

    PCells = sheet[ProjectCell]
    NCells = sheet[NoteCell]
    PMerge = ProjectCell + ':I' + str(PCells.row)
    NMerge = NoteCell + ':I' + str(NCells.row)
    
#   Set the alignement parameter    
    Al = Alignment(horizontal='left',
                   vertical='top',
                   wrap_text=True)
#   Apply the alignement to the first cell of the Note range
    PCells.alignment = Al
    NCells.alignment = Al

#   Make the cell height acceptable to accomodate multiple lines
    sheet.row_dimensions[NCells.row].height =250
#   Merge Note cells to allow fot them to be all displayed
    sheet.merge_cells(PMerge)
    sheet.merge_cells(NMerge)

    sheet.column_dimensions['A'].width = 26
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 10
    
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    sheet.oddHeader.left.text = 'Project: ' + str(Project)

    wb.save(FileName)